import pandas as pd
import numpy as np
import torch
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error
import dill
from tqdm import tqdm
import openpyxl
from utils.metrics import get_average_metrics
import scipy.stats as stats

import json

class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.float32):
            # 将 np.float32 转换为 float 类型并保留 4 位小数
            return round(float(obj), 4)
        return super().default(obj)
def save_dict_to_txt(data_dict, file_path):
    """
    将字典数据以每行一个字典的形式保存到 TXT 文件中，支持多次写入不覆盖
    :param data_dict: 包含数据的字典
    :param file_path: 保存文件的路径
    """
    with open(file_path, 'a', encoding='utf-8') as txtfile:
        # 将字典转换为 JSON 字符串
        json_str = json.dumps(data_dict,cls=NumpyEncoder)
        # 写入一行字典的 JSON 字符串
        txtfile.write(json_str + '\n')

import csv
# def metrics_of_pv(preds, trues):
#     pred = np.array(preds)
#     true = np.array(trues)
#     mae = np.round(mean_absolute_error(true, pred),4)
#     rmse = np.round(np.sqrt(mean_squared_error(true, pred)),4)
#     r2 = np.round(r2_score(true, pred),4)
#     mbe = np.round(np.mean(pred - true),4)
#     # sMAPE = np.round(100 * np.mean(np.abs(preds - trues) / (np.abs(preds) + np.abs(trues))), 4)
#     # mape = np.round(100 * np.mean(np.abs((true - pred) / true)), 4)
#     return {'MAE': mae, 'RMSE': rmse, 'R2': r2, 'MBE': mbe,}

def metrics_of_pv(preds, trues, reduce_mean=True):
    """
    Compute evaluation metrics for photovoltaic predictions.

    Parameters:
    preds (np.ndarray): Predicted values, shape (b, l).
    trues (np.ndarray): True values, shape (b, l).
    reduce_mean (bool): Whether to average metrics across all time steps.

    Returns:
    dict: Metrics for each prediction step or averaged across steps.
    """
    preds = np.array(preds)
    trues = np.array(trues)

    # Compute metrics for each prediction step
    mae = np.round(np.mean(np.abs(trues - preds), axis=0), 4)
    rmse = np.round(np.sqrt(np.mean((trues - preds) ** 2, axis=0)), 4)
    r2 = np.round(np.array([r2_score(trues[:, i], preds[:, i]) for i in range(preds.shape[1])]), 4)
    mbe = np.round(np.mean(preds - trues, axis=0), 4)

    if reduce_mean:
        return {
            'MAE': np.round(np.mean(mae), 4),
            'RMSE': np.round(np.mean(rmse), 4),
            'R2': np.round(np.mean(r2), 4),
            'MBE': np.round(np.mean(mbe), 4),
        }
    else:
        return {
            'MAE': mae.tolist(),
            'RMSE': rmse.tolist(),
            'R2': r2.tolist(),
            'MBE': mbe.tolist(),
        }

def metrics_v1(preds, trues):
    preds = np.array(preds)
    trues = np.array(trues)
    MAE, RMSE, R2, MBE, MAPE = [], [], [], [], []
    preds=np.array_split(preds,3,-1)
    trues=np.array_split(trues,3,-1)
    for i in range(3):
        true=trues[i]
        print(np.min(true),np.max(true))
        pred=preds[i]
        mae = np.round(mean_absolute_error(true, pred),4)
        rmse = np.round(np.sqrt(mean_squared_error(true, pred)),4)
        r2 = np.round(r2_score(true, pred),4)
        mbe = np.round(np.mean(pred - true),4)
        # sMAPE = np.round(100 * np.mean(np.abs(preds - trues) / (np.abs(preds) + np.abs(trues))), 4)
        mape = np.round(100 * np.mean(np.abs((true - pred) / (true+0.01))), 4)
        MAE.append(mae)
        RMSE.append(rmse)
        R2.append(r2)
        MAPE.append(mape)
    return {'MAE': MAE, 'RMSE': RMSE, 'R2': R2, 'MAPE': MAPE}

def save2csv(n, file):
    n = n.reshape((1, n.shape[0]))
    n = pd.DataFrame(n)
    n.to_csv(file, index=False, encoding='utf-8', header=False, mode='a')

# def same_seeds(seed):
#     torch.manual_seed(seed)  # 固定随机种子（CPU）
#     if torch.cuda.is_available():  # 固定随机种子（GPU)
#         torch.cuda.manual_seed(seed)  # 为当前GPU设置
#         torch.cuda.manual_seed_all(seed)  # 为所有GPU设置
#     np.random.seed(seed)  # 保证后续使用random函数时，产生固定的随机数
#     torch.backends.cudnn.benchmark = True  # GPU、网络结构固定，可设置为True
#     # torch.backends.cudnn.deterministic = True  # 固定网络结构

# def same_seeds(seed):
#     np.random.seed(seed)
#     np.random.seed(seed)
#     torch.manual_seed(seed)
#     torch.cuda.manual_seed(seed)
#     torch.cuda.manual_seed_all(seed)
#     # torch.backends.cudnn.deterministic = True
#     torch.backends.cudnn.benchmark = False
def same_seeds(seed):
    torch.manual_seed(seed)  # 固定随机种子（CPU）
    if torch.cuda.is_available():  # 固定随机种子（GPU)
        torch.cuda.manual_seed(seed)  # 为当前GPU设置
        torch.cuda.manual_seed_all(seed)  # 为所有GPU设置
    np.random.seed(seed)  # 保证后续使用random函数时，产生固定的随机数
    torch.backends.cudnn.benchmark = True  # GPU、网络结构固定，可设置为True
    # torch.backends.cudnn.deterministic = True  # 固定网络结构

def save_dict_to_excel(data_dict, file_name='model_results.xlsx', sheet_name='Sheet1'):
    # 尝试打开现有的工作簿，如果文件不存在则创建一个新的
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    # 如果指定的工作表不存在，则创建新的
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]
    # 获取所有字段（列名），确保第一列是 model_name
    fieldnames = ['model_name'] + [key for key in data_dict.keys() if key != 'model_name']
    # 判断表头是否已经存在
    if sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None:
        # 如果表头不存在，则写入表头到第一行
        for col_num, fieldname in enumerate(fieldnames, start=1):
            sheet.cell(row=1, column=col_num, value=fieldname)
    # 找到下一行
    next_row = sheet.max_row + 1
    # 构建行数据，确保 model_name 在第一列
    row = {'model_name': data_dict['model_name'],
           **{key: data_dict.get(key, None) for key in fieldnames if key != 'model_name'}}
    # 将数据写入到下一行
    for col_num, fieldname in enumerate(fieldnames, start=1):
        sheet.cell(row=next_row, column=col_num, value=row[fieldname])
    # 保存文件
    workbook.save(file_name)


class EarlyStopping:
    """Early stops the training if validation loss doesn't improve after a given patience."""

    def __init__(self, save_path, patience=7, verbose=False, delta=0):
        """
        Args:
            save_path : 模型保存文件夹
            patience (int): How long to wait after last time validation loss improved.
                            Default: 7
            verbose (bool): If True, prints a message for each validation loss improvement.
                            Default: False
            delta (float): Minimum change in the monitored quantity to qualify as an improvement.
                            Default: 0
        """
        self.save_path = save_path
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.delta = delta

    def __call__(self, val_loss, model,hyperparameters=None):

        score = -val_loss

        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model,hyperparameters)
        elif score < self.best_score + self.delta:
            self.counter += 1
            print(f"EarlyStopping counter: {self.counter} out of {self.patience}")
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model,hyperparameters)
            self.counter = 0

    def save_checkpoint(self, val_loss, model,hyperparameters=None):
        """Saves model when validation loss decrease."""
        if self.verbose:
            print(
                f"Validation loss decreased ({self.val_loss_min:.6f} --> {val_loss:.6f}).  Saving model ..."
            )
        # path = os.path.join(self.save_path, 'best_network.pth')
        path = self.save_path
        checkpoint = {
                'model_state_dict': model.state_dict(),
                'hyperparameters': hyperparameters,
            }
        torch.save(checkpoint, path)  # 这里会存储迄今最优模型的参数
        self.val_loss_min = val_loss

def train(data, model, criterion, optm, device=torch.device("cuda:0")):
    model.train()
    running_loss = 0.0
    for x, y in tqdm(data):
        model.zero_grad()
        x, y = x.float().to(device), y.float().to(device)
        optm.zero_grad()
        y_pre = model(x)
        loss = criterion(y_pre, y)
        loss.backward()
        optm.step()
        running_loss += loss.item() * x.size(0)
    epoch_loss = running_loss / len(data.dataset)
    return epoch_loss
def evaluate(data, model, criterion, device=torch.device("cuda:0"),scalar=None,prob=False,v1=False,prob_type=None):
    model.eval()
    val_running_loss = 0.0
    # all_preds = []
    # all_labels = []
    all_preds = None
    all_labels = None
    for x, y in tqdm(data):
        model.zero_grad()
        with torch.no_grad():
            x, y = x.float().to(device), y.float().to(device)
            y_pre = model(x)
            loss = criterion(y_pre, y)
            val_running_loss += loss.item() * x.size(0)
            if all_preds is None:
                all_preds = y_pre.cpu().numpy()
                all_labels = y.cpu().numpy()
            else:
                all_preds = np.concatenate((all_preds, y_pre.cpu().numpy()), axis=0)
                all_labels = np.concatenate((all_labels, y.cpu().numpy()), axis=0)
            # all_preds.extend(y_pre.cpu().numpy())
            # all_labels.extend(y.cpu().numpy())
    epoch_loss = val_running_loss / len(data.dataset)
    if prob_type == 'Gauss' or prob_type == 'Laplace':
        _, p, l = all_preds.shape
        mu = all_preds[:, :, 0]
        sigma = all_preds[:, :, 1]
        preds_ = np.zeros((all_preds.shape[0], p, 7))
        preds_[:, :, 0] = mu
        for i, level in enumerate([0.95, 0.90, 0.85]):
            i = i + 1
            if prob_type == 'Gauss':
                preds_[:, :, i], preds_[:, :, -i] = stats.norm.interval(level, loc=mu, scale=sigma)
            elif prob_type == 'Laplace':
                preds_[:, :, i], preds_[:, :, -i] = stats.laplace.interval(level, loc=mu, scale=sigma)
        all_preds = preds_
    elif prob_type == 'SDER':
        _, p, l = all_preds.shape
        preds_ = np.zeros((all_preds.shape[0], p, 7))
        gamma, nu, alpha, beta = all_preds[:, :, 0], all_preds[:, :, 1], all_preds[:, :, 2], all_preds[:, :, 3],
        preds_[:, :, 0] =gamma
        # sigma = np.sqrt(beta / (alpha - 1))  # Aleatoric
        # mu = np.sqrt(beta / (nu * (alpha - 1)))  # Epistemic
        student_var = np.sqrt(beta * (1. + nu) / (nu * alpha))
        for i, level in enumerate([0.95, 0.90, 0.85]):
            i = i + 1
            preds_[:, :, i], preds_[:, :, -i] = stats.t.interval(level, loc=gamma, scale=student_var, df=2 * alpha)
        all_preds = preds_

    if scalar is not None:
        # 点预测反归一化
        if len(all_preds.shape) == 2:
            all_preds=scalar.inverse_transform(all_preds)
            all_labels = scalar.inverse_transform(all_labels)
        # 区间预测反归一化
        elif len(all_preds.shape) == 3:
            all_labels = scalar.inverse_transform(all_labels)
            _, p, l = all_preds.shape
            all_preds = scalar.inverse_transform(all_preds.reshape(all_preds.shape[0], -1)).reshape(-1, p, l)
        else:
            raise ValueError('The shape of all_preds is wrong')
        all_preds = np.maximum(all_preds, 0)
        all_labels = np.maximum(all_labels, 0)
    if not prob:
        if not v1:
            metrics_ = metrics_of_pv(all_preds, all_labels,reduce_mean=True)
        else:

            metrics_= metrics_v1(all_preds, all_labels)
    else:

        metrics_= get_average_metrics(all_preds, all_labels)
    return epoch_loss, metrics_,[all_preds,all_labels]


import torch
import torch.nn.functional as F
import matplotlib.pyplot as plt
import numpy as np

import torch
import torch.nn.functional as F
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False

def cam_analysis(model, input_data, layer_name='conv1'):
    """
    对回归任务的每个时间步进行 CAM 分析，并将每个时间步的 CAM 绘制成 4x1 图。

    :param model: 已训练的 PyTorch 模型
    :param input_data: 输入数据，形状为 (1, 72)
    :param layer_name: 要分析的卷积层名称
    :return: 叠加后的时序数据与 CAM 的结果
    """

    # 注册钩子获取指定层的输出和梯度
    def register_hooks(model, layer_name):
        def hook_fn(module, input, output):
            model.activations = output  # 保存该层的激活值
            def backward_hook(module, grad_input, grad_output):
                model.gradients = grad_output[0]  # 保存该层的梯度
            module.register_backward_hook(backward_hook)

        layer = dict(model.named_modules())[layer_name]
        layer.register_forward_hook(hook_fn)

    # 注册钩子函数
    register_hooks(model, layer_name)
    device=torch.device('cuda:0')
    # 转换输入数据为 Tensor，并进行前向传播
    # input_data = torch.tensor(input_data, dtype=torch.float32).unsqueeze(0)  # shape: (1, 72)
    input_data=input_data.float().to(device)
    input_data.requires_grad_()
    output = model(input_data)  # 假设模型输出为 (1, 4)

    # 创建 4x1 的子图布局
    fig, axs = plt.subplots(4, 1, figsize=(10, 12))

    # 遍历每个时间步并计算 CAM
    for time_step in range(4):  # 假设你要分析预测的 4 个时间步
        # 对每个时间步进行反向传播
        model.zero_grad()
        output[:, time_step].sum().backward(retain_graph=True)  # 对目标时间步执行反向传播

        # 获取模型的激活图（这里假设是通过钩子获取）
        activations = model.activations  # 这是一个假设，视你的模型而定

        # 获取输入数据的梯度
        gradients = input_data.grad  # 获取输入数据的梯度

        # 对每个通道的梯度进行全局平均池化（Global Average Pooling）
        weights = torch.mean(gradients[:,:,0], dim=(0, 1))  # 对所有 batch 和序列长度求均值

        # 计算 CAM 图（在时序数据上，CAM 代表每个时间步的权重）
        cam = torch.zeros(activations.shape[2:], dtype=torch.float32).to(device)  # 修改这里适配新的 activations 维度
        # for i in range(activations.shape[2]):  # 遍历特征维度
        cam += weights * activations[0, 0, :]  # 对每个特征维度进行加权和

        # 对 CAM 进行 ReLU 激活
        cam = F.relu(cam)

        # 归一化 CAM 图
        cam = cam - torch.min(cam)
        cam = cam / torch.max(cam)

        # 转换为 numpy 数组方便可视化
        cam = cam.detach().cpu().numpy()

        # 在子图中绘制该时间步的 CAM
        axs[time_step].plot(cam, label=f'Time Step {time_step + 1}')
        axs[time_step].set_xlabel('Time Step')
        axs[time_step].set_ylabel('Activation')
        axs[time_step].set_title(f'CAM for Time Step {time_step + 1}')
        axs[time_step].legend()

        # 调整布局
    plt.tight_layout()
    plt.show()

    # # 归一化原始时序数据（input_data 是 (300, 72, 5)）
    # original_sequence = input_data[0, :, 0].detach().cpu().numpy()  # 假设我们只取第一个特征的序列
    # original_sequence = (original_sequence - np.min(original_sequence)) / (
    #             np.max(original_sequence) - np.min(original_sequence))
    #
    # # 将 CAM 图与原始时序数据叠加
    # cam_resized = np.interp(cam, (np.min(cam), np.max(cam)), (0, 1))  # CAM 归一化
    # overlay = 0.7 * original_sequence + 0.3 * cam_resized
    #
    # # 可视化叠加后的时序数据与 CAM
    # plt.plot(overlay, label="Overlayed Data with CAM")
    # plt.plot(original_sequence, label="Original Data")
    # plt.legend()
    # plt.title("Time Series with CAM Overlay")
    # plt.show()
    return None
    # return overlay  # 返回叠加后的时序数据和 CAM


def visualize_feature_map(model, input_data, layer_name='conv1', downsample_factor=8):
    """
    可视化指定层的特征图（矩形块形式）并适当下采样。

    :param model: 已训练的 PyTorch 模型
    :param input_data: 输入数据，形状为 (1, sequence_length, num_features)
    :param layer_name: 要可视化的层的名称
    :param downsample_factor: 下采样因子，控制矩形块的大小
    """
    activations = []

    # 注册钩子获取指定层的输出
    def hook_fn(module, input, output):
        activations.append(output)

    # 获取目标层并注册钩子
    layer = dict(model.named_modules())[layer_name]
    layer.register_forward_hook(hook_fn)

    # 转换输入数据为 Tensor 并执行前向传播
    input_data = input_data.float().to(next(model.parameters()).device)  # 确保输入数据和模型在同一设备上
    with torch.no_grad():
        _ = model(input_data)

    # 提取激活值（特征图）
    feature_map = activations[0]  # 维度为 (batch_size, 1, num_features)

    # 假设 batch_size 为 1，提取特征图并转换为 NumPy
    feature_map = feature_map[:10].detach().cpu().numpy()  # 维度为 (512,)

    # 对特征图进行下采样（降低分辨率）
    if downsample_factor > 1:
        # feature_map = feature_map.reshape(1, 1, feature_map.shape[0])  # 重新调整维度为 (1, 1, 512)
        feature_map = F.interpolate(torch.tensor(feature_map), scale_factor=1 / downsample_factor, mode='linear', align_corners=False)
        feature_map = feature_map.squeeze(0).squeeze(0).detach().cpu().numpy()  # 恢复为 (downsampled_length,)
    feature_map = (feature_map - np.min(feature_map)) / (np.max(feature_map) - np.min(feature_map) + 1e-8)
    print(feature_map.shape)
    np.save('data/feature_map_CA.npy', feature_map)

    ksize=20
    # 绘制特征图（矩形块形式）
    plt.figure(figsize=(10, 6))
    im=plt.imshow(feature_map.reshape(10, -1), aspect=6, cmap="viridis", interpolation="none")
    cbar = plt.colorbar(im)
    # 设置颜色条标签的字体大小
    cbar.set_label('Values', fontsize=ksize)
    # 设置颜色条刻度的字体大小
    cbar.ax.tick_params(labelsize=ksize)
    # plt.colorbar(label='特征值',fontsize=16)
    plt.xlabel("Index", fontsize=ksize)
    # plt.ylabel("通道")
    plt.title(f"Feature map of GTU", fontsize=ksize)
    # plt.title(f"CA特征图", fontsize=ksize)
    plt.yticks([0, 5, 9], ['1', '6', '10'])
    plt.tick_params(axis='x', labelsize=ksize)
    plt.tick_params(axis='y', labelsize=ksize)

    plt.tight_layout()
    # plt.savefig('./pic/feature_map_CA.svg', format='SVG', dpi=800)
    plt.savefig('./pic/feature_map_GTU.svg', format='SVG', dpi=800)
    plt.show()






