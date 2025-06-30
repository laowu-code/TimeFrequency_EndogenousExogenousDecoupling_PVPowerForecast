# def select_hyperparameters(trial, model_name='Proposed', type='single', seq_len=48, pred_len=4, enc_in=5):
#     params_base = {'seq_len': seq_len, 'pred_len': pred_len, 'enc_in': enc_in, }
#     if type == 'single':
#         if model_name == 'Proposed':
#             params = {'d_model': 256,  'n_heads': 6,'e_layers': 3, "embed_size": 128, 'hidden_size': 256}
#         elif model_name == 'Dlinear':
#             params = {'moving_avg': 4, "individual": False}
#         elif model_name == 'FreTS':
#             params = {'embed_size': 128, 'hidden_size': 256}
#         elif model_name == 'LSTM':
#             params = {'d_model': 128, 'e_layers': 1}
#         elif model_name == 'GRU':
#             params = {'d_model': 128, 'e_layers': 1, }
#         elif model_name == 'TCN':
#             params = {'channels': 128, 'e_layers': 3, 'kernel_size': 3, }
#         elif model_name == 'Pyraformer':
#             params = {'d_ff': 128, 'n_heads': 4, 'e_layers': 1, 'd_model': 128, }
#         elif model_name == 'iTransformer':
#             params = {'d_model': 128, 'e_layers': 2, 'n_heads': 4,}
#         elif model_name == 'PatchTST':
#             params = {'d_model': 256,  'n_heads': 6, 'e_layers': 3, 'patch_len': 4, 'stride_flag': 'full'}
#         elif model_name == 'TimeXer':
#             params = {'d_model': 256, 'enc_in': 1, 'n_heads': 6, 'e_layers': 3, 'patch_len': 4}
#         else:
#             raise ValueError('Model name not found')
#     elif type == 'optuna':
#         if model_name == 'Proposed':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'n_heads': trial.suggest_categorical('n_heads', [2, 4, 6, 8, 12]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       "embed_size": trial.suggest_categorical('embed_size', [16, 32, 64, 128, 256, 512]),
#                       'hidden_size': trial.suggest_categorical('hidden_size', [16, 32, 64, 128, 256, 512])}
#         elif model_name == 'Dlinear':
#             params = {'moving_avg': trial.suggest_categorical('moving_avg', [2, 4, 6, 8, 12]),
#                       "individual": trial.suggest_categorical('individual', [True, False])}
#         elif model_name == 'FreTS':
#             params = {'embed_size': trial.suggest_categorical('embed_size', [16, 32, 64, 128, 256, 512]),
#                       'hidden_size': trial.suggest_categorical('hidden_size', [16, 32, 64, 128, 256, 512])}
#         elif model_name == 'LSTM':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6])}
#         elif model_name == 'GRU':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6])}
#         elif model_name == 'TCN':
#             params = {'channels': trial.suggest_categorical('channels', [16, 32, 64, 128, 256, 512]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       'kernel_size': trial.suggest_categorical('kernel_size', [2, 3, 4, 5, 6, 7, 8])}
#         elif model_name == 'Pyraformer':
#             params = {
#                       'n_heads': trial.suggest_categorical('n_heads', [2, 4, 6, 8, 12]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       'd_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512])}
#         elif model_name == 'iTransformer':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       'n_heads': trial.suggest_categorical('n_heads', [2, 4, 6, 8, 12]),}
#         elif model_name == 'PatchTST':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'n_heads': trial.suggest_categorical('n_heads', [2, 4, 6, 8, 12]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       'patch_len': trial.suggest_categorical('patch_len', [2, 3, 4, 5, 6, 7, 8]),
#                       'stride_flag': trial.suggest_categorical('stride_flag', ['full', 'half'])}
#         elif model_name == 'TimeXer':
#             params = {'d_model': trial.suggest_categorical('d_model', [16, 32, 64, 128, 256, 512]),
#                       'n_heads': trial.suggest_categorical('n_heads', [2, 4, 6, 8, 12]),
#                       'e_layers': trial.suggest_categorical('e_layers', [1, 2, 3, 4, 5, 6]),
#                       'patch_len': trial.suggest_categorical('patch_len', [2, 3, 4, 5, 6, 7, 8])}
#         else:
#             raise ValueError('Model name not found')
#     elif type == 'optimal':
#         pass
#     else:
#         raise ValueError('Type not found')
#     params={**params_base, **params}
#     return params
#
# print(select_hyperparameters(1, model_name='TimeXer', type='single',))
import openpyxl


# def append_dict_to_excel(data_dict, file_name='model_results.xlsx', sheet_name='Sheet1'):
#     # 尝试打开现有的工作簿，如果文件不存在则创建一个新的
#     try:
#         workbook = openpyxl.load_workbook(file_name)
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#
#     # 如果指定的工作表不存在，则创建新的
#     if sheet_name not in workbook.sheetnames:
#         workbook.create_sheet(sheet_name)
#
#     sheet = workbook[sheet_name]
#
#     # 获取所有字段（列名），确保第一列是 model_name
#     fieldnames = ['model_name'] + [key for key in data_dict.keys() if key != 'model_name']
#
#     # 判断表头是否已经存在
#     if sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None:
#         # 如果表头不存在，则写入表头到第一行
#         for col_num, fieldname in enumerate(fieldnames, start=1):
#             sheet.cell(row=1, column=col_num, value=fieldname)
#
#     # 找到下一行
#     next_row = sheet.max_row + 1
#
#     # 构建行数据，确保 model_name 在第一列
#     row = {'model_name': data_dict['model_name'],
#            **{key: data_dict.get(key, None) for key in fieldnames if key != 'model_name'}}
#
#     # 将数据写入到下一行
#     for col_num, fieldname in enumerate(fieldnames, start=1):
#         sheet.cell(row=next_row, column=col_num, value=row[fieldname])
#
#     # 保存文件
#     workbook.save(file_name)



# 示例用法：
for idx in range(1,10):
    print(idx)

# 示例用法：
# data = {'MAE': 0.1029, 'RMSE': 0.2112, 'R2': 0.9479, 'MBE': -0.0101, 'model_name': 'FGT'}

